import sys

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import *
import uuid
from openpyxl import Workbook, load_workbook

class tab_demo(QMainWindow):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.data_file = "task_data.xlsx"  # Excel数据文件路径
        self.resize(1500, 1000)  # 设置窗口初始大小
        self.setWindowTitle("小红书获客系统v1.00")  # 设置窗口标题

        # 创建选项卡控件
        self.tabs = QTabWidget()
        self.setCentralWidget(self.tabs)  # 将选项卡设置为主窗口的中心部件

        # 创建两个标签页
        self.tab1 = QWidget()
        self.tab2 = QWidget()

        # 添加标签页并设置名称
        self.tabs.addTab(self.tab1, "关键词追踪任务")
        self.tabs.addTab(self.tab2, "帖子ID追踪任务")

        # 初始化两个标签页的UI
        self.tab1UI()
        self.tab2UI()
        # 初始化时加载历史数据
        self.load_data()

    def load_data(self):
        """从Excel文件加载历史数据"""
        try:
            wb = load_workbook(self.data_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
                uuid_val, city, keyword, post_sort, analyze_comments,track_count, status = row
                self.add_row(uuid_val,city, keyword, post_sort, analyze_comments, track_count, status)
        except FileNotFoundError:
            # 如果文件不存在，创建新的Excel文件
            wb = Workbook()
            ws = wb.active
            ws.append(["UUID", "城市", "关键词", "帖子排序", "分析评论", "跟踪帖子数量", "状态"])
            wb.save(self.data_file)

    def save_to_excel(self, UUID,city, keyword, post_sort, analyze_comments, track_count, status):
        """保存数据到Excel文件"""
        wb = load_workbook(self.data_file)
        ws = wb.active
        ws.append([
            UUID,  # 生成UUID
            city,
            keyword,
            post_sort,
            analyze_comments,
            track_count,
            status
        ])
        wb.save(self.data_file)
    def tab1UI(self):
        """初始化第一个标签页的UI布局"""
        layout = QVBoxLayout()  # 使用垂直布局

        # 创建并设置"添加"按钮
        self.add_button = QPushButton("添加")
        self.add_button.setFixedSize(50, 50)  # 固定按钮大小
        add_button_layout = QHBoxLayout()
        add_button_layout.addStretch()  # 添加弹性空间使按钮靠右
        add_button_layout.addWidget(self.add_button)
        self.add_button.clicked.connect(self.open_form)  # 绑定点击事件
        layout.addLayout(add_button_layout)

        # 创建表格控件
        self.table = QTableWidget(0, 8)  # 初始0行7列
        self.table.setHorizontalHeaderLabels(["UUID","城市", "关键词", "帖子排序", "分析评论", "跟踪帖子数量", "状态", "操作"])
        self.table.horizontalHeader().setDefaultAlignment(Qt.AlignLeft)  # 表头左对齐
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)  # 禁止编辑表格内容
        layout.addWidget(self.table)  # 将表格添加到布局

        self.tab1.setLayout(layout)  # 设置标签页的布局

    def add_row(self,UUID, city, keyword, post_sort, analyze_comments, track_count, status):
        """向表格添加新行"""
        row_position = self.table.rowCount()
        self.table.insertRow(row_position)  # 插入新行

        # 添加前5列的数据
        for col, value in enumerate([UUID,city, keyword, post_sort, analyze_comments, track_count, status]):
            item = QTableWidgetItem(value)
            item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)  # 设置单元格文本对齐方式
            item.setFlags(item.flags() & ~Qt.ItemIsEditable)  # 设置单元格不可编辑
            self.table.setItem(row_position, col, item)

        # 创建操作按钮布局
        btn_layout = QHBoxLayout()

        # 创建四个操作按钮
        restart_btn = QPushButton("重启")
        # edit_btn = QPushButton("编辑")
        # delete_btn = QPushButton("删除")

        # 统一设置按钮大小
        for btn in [restart_btn]:
            btn.setFixedSize(50, 50)

        # 添加按钮到布局，并设置间距
        btn_layout.addStretch()  # 左侧弹性空间
        btn_layout.addWidget(restart_btn)
        btn_layout.addSpacing(10)  # 按钮间距
        # btn_layout.addWidget(edit_btn)
        # btn_layout.addSpacing(10)
        # btn_layout.addWidget(delete_btn)
        btn_layout.addStretch()  # 右侧弹性空间
        btn_layout.setAlignment(Qt.AlignLeft)
        # 绑定按钮事件
        # edit_btn.clicked.connect(lambda: self.open_form(edit=True, row=row_position))
        # delete_btn.clicked.connect(lambda: self.table.removeRow(row_position))

        # 将按钮布局放入容器并添加到表格单元格
        container = QWidget()
        container.setLayout(btn_layout)
        self.table.setCellWidget(row_position, 7, container)  # 添加到第6列(操作列)

    def open_form(self, edit=False, row=None):
        if edit and row is not None:
            print(f"编辑任务，行号: {row}")
            return  # 如果是编辑模式但没有指定行，不执行操作
        """打开添加/编辑任务的对话框"""
        dialog = QDialog(self)
        dialog.setWindowTitle("添加/编辑任务")
        form_layout = QFormLayout()  # 使用表单布局

        # 创建输入控件
        city_input = QLineEdit()
        keyword_input = QLineEdit()

        # 创建帖子排序单选按钮组
        group = QButtonGroup()
        post_sort_layout = QHBoxLayout()
        r1 = QRadioButton("最新")
        r2 = QRadioButton("最热")
        r3 = QRadioButton("综合")
        group.addButton(r1)
        group.addButton(r2)
        group.addButton(r3)
        post_sort_layout.addWidget(r1)
        post_sort_layout.addWidget(r2)
        post_sort_layout.addWidget(r3)
        post_sort_layout.addStretch()  # 添加伸缩空间，让单选按钮靠左

        # 创建跟踪数量单选按钮组
        num_group = QButtonGroup()

        track_count_layout = QHBoxLayout()
        ten_checkbox = QRadioButton("10")
        twenty_checkbox = QRadioButton("20")
        thirty_checkbox = QRadioButton("30")
        num_group.addButton(ten_checkbox)
        num_group.addButton(twenty_checkbox)
        num_group.addButton(thirty_checkbox)
        track_count_layout.addWidget(ten_checkbox)
        track_count_layout.addWidget(twenty_checkbox)
        track_count_layout.addWidget(thirty_checkbox)

        # 创建是否分析评论复选框
        analyze_comments_group = QButtonGroup()

        analyze_comments_layout = QHBoxLayout()
        analyze_comments_yes = QRadioButton("是")
        analyze_comments_no = QRadioButton("否")
        analyze_comments_group.addButton(analyze_comments_yes)
        analyze_comments_group.addButton(analyze_comments_no)
        analyze_comments_layout.addWidget(analyze_comments_yes)
        analyze_comments_layout.addWidget(analyze_comments_no)
        # 添加表单行
        form_layout.addRow("城市", city_input)
        form_layout.addRow("关键词", keyword_input)
        form_layout.addRow("帖子排序", post_sort_layout)
        form_layout.addRow("跟踪帖子数量", track_count_layout)
        form_layout.addRow("是否分析评论", analyze_comments_layout)

        status = '运行中'
        # 如果是编辑模式，填充现有数据
        if edit and row is not None:
            city_input.setText(self.table.item(row, 0).text())
            keyword_input.setText(self.table.item(row, 1).text())

        # 创建保存按钮
        save_button = QPushButton("保存")
        save_button.clicked.connect(
            lambda: self.save_form(dialog, city_input.text(), keyword_input.text(),
                                   "最新" if r1.isChecked() else "最热" if r2.isChecked() else "综合" if r3.isChecked() else "",
                                   "是" if analyze_comments_yes.isChecked() else "否" if analyze_comments_no.isChecked() else "",
                                   "10" if ten_checkbox.isChecked() else "20" if twenty_checkbox.isChecked() else "30" if thirty_checkbox.isChecked() else "",
                                   status))
        save_button.setFixedSize(50, 50)
        save_button_layout = QHBoxLayout()
        save_button_layout.addStretch()
        save_button_layout.addWidget(save_button)
        save_button_layout.addStretch()
        form_layout.addRow(save_button_layout)

        dialog.setLayout(form_layout)
        dialog.exec_()  # 显示对话框

    def save_form(self, dialog, city, keyword, post_sort, analyze_comments, track_count, status):
        """保存表单数据并关闭对话框"""
        uuid_str = str(uuid.uuid4())
        print(f"Saved:uuid_str={uuid_str} City={city}, Keyword={keyword}, Post Sort={post_sort}, Analyze Comments={analyze_comments}, track_count={track_count},status={status}")
        self.add_row(uuid_str, city, keyword, post_sort, analyze_comments, track_count, status)
        self.save_to_excel(uuid_str ,city, keyword, post_sort, analyze_comments,track_count, status)

        # 启动任务，
        dialog.accept()  # 关闭对话框

    def tab2UI(self):
        """初始化第二个标签页的UI布局"""
        layout = QFormLayout()
        self.tab2.setLayout(layout)


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = tab_demo()
    ex.show()
    sys.exit(app.exec_())
